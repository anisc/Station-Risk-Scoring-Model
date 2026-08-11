#!/usr/bin/env python3
"""Process uploaded data files and regenerate JS data files."""
import csv, json, os, re

ROOT = os.path.dirname(os.path.abspath(__file__))


def find_data_file(basename):
    """Find a data file by base name, trying .csv then .xlsx."""
    for ext in ['.csv', '.xlsx']:
        p = os.path.join(ROOT, basename + ext)
        if os.path.exists(p):
            return p
    return None


def read_data_rows(filepath):
    """Read rows from an xlsx or csv file. Returns list of dicts."""
    if not filepath or not os.path.exists(filepath):
        return []
    if filepath.lower().endswith('.csv'):
        with open(filepath, 'r', encoding='utf-8-sig') as f:
            return [{k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items()} for row in csv.DictReader(f)]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not all_rows:
            return []
        header = [str(h).strip() if h else '' for h in all_rows[0]]
        rows = []
        for r in all_rows[1:]:
            row = {}
            for i, h in enumerate(header):
                if i < len(r) and r[i] is not None:
                    val = r[i]
                    if isinstance(val, str):
                        val = val.strip()
                    elif hasattr(val, 'strftime'):
                        val = val.strftime('%Y-%m-%d')
                    else:
                        val = str(val).strip() if val else ''
                    row[h] = val
            if any(row.values()):
                rows.append(row)
        return rows

# Region mapping
REGION_ICAO = {
    'Western Canada & Mexico': ['C', 'MM', 'MY', 'MZ'],
    'Central Canada & LATAM': ['SB', 'SP', 'SC', 'SE', 'SU', 'SK'],
    'Eastern Canada & Europe & Asia': ['CY', 'CZ', 'EG', 'EI', 'LF', 'ED', 'EH', 'EK', 'LP', 'LE', 'LI', 'RJ', 'RC', 'RK', 'VT', 'VD', 'WB', 'GC', 'GE', 'GV', 'GO', 'GA', 'GF', 'GL', 'DT', 'TS', 'GMM', 'GM', 'GN'],
    'US & Caribbean': ['K', 'TJ', 'TN', 'TR', 'TI', 'TD', 'TK', 'TP', 'TA', 'TL', 'TG', 'MK', 'MD', 'SJ', 'SI', 'NS', 'PA'],
}

def get_region_from_icao(icao):
    """Determine region from an ICAO code using prefix matching."""
    if not icao or len(icao) < 2:
        return None
    icao = icao.upper()
    for region, prefixes in REGION_ICAO.items():
        for prefix in prefixes:
            if icao.startswith(prefix):
                return region
    return None


def get_station_region(iata):
    """Determine region from ICAO prefix pattern on IATA."""
    return None  # Region comes from the CRS data directly

def normalize_header(h):
    """Normalize header names across CSV and XLSX formats."""
    h = h.strip()
    mapping = {
        'Report Created': 'ReportCreated',
        'Report No': 'ReportNo',
        'Import Date': 'ImportDate',
        'Imported By': 'ImportedBy',
        'Occ No': 'OccNo',
        'Report Desc': 'ReportDesc',
        'Occ Desc': 'OccDesc',
        'Incident City': 'IncidentCity',
        'Departure City': 'DepartureCity',
        'Destination City': 'DestinationCity',
        'OAPT Concerns': 'OAPT_Concerns',
        'Date Closed': 'DateClosed',
        'Closed By': 'ClosedBy',
    }
    return mapping.get(h, h)


def process_oapt():
    """Process OAPT Occurrence Report into oapt_reports.js and help CRS merge."""
    xlsx_path = find_data_file('OAPT Occurrence Report')
    if xlsx_path and not xlsx_path.endswith('.csv'):
        pass  # found xlsx
    else:
        xlsx_path = os.path.join(ROOT, 'OAPT Occurrence Report.xlsx')  # fallback
    csv_path = None
    for name in ['OAPT Occurrence Report.csv', 'OAPT Occurrence Report(1).csv']:
        p = os.path.join(ROOT, name)
        if os.path.exists(p):
            csv_path = p
            break

    rows = []
    if os.path.exists(xlsx_path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            all_rows = list(ws.iter_rows(values_only=True))
            if all_rows:
                # Build header mapping from first row, skipping None columns
                raw_header = all_rows[0]
                header = []
                col_indices = []
                for i, h in enumerate(raw_header):
                    if h is not None and str(h).strip():
                        header.append(normalize_header(str(h).strip()))
                        col_indices.append(i)

                for r in all_rows[1:]:
                    row = {}
                    for j, h in enumerate(header):
                        idx = col_indices[j]
                        if idx < len(r) and r[idx] is not None:
                            val = r[idx]
                            if isinstance(val, str):
                                val = val.strip()
                            elif hasattr(val, 'strftime'):
                                val = val.strftime('%Y-%m-%d')
                            else:
                                val = str(val).strip() if val else ''
                            row[h] = val
                    if row.get('OccNo'):
                        rows.append(row)
            wb.close()
        except Exception as e:
            print(f"Warning: Could not read OAPT xlsx: {e}")

    if not rows and os.path.exists(csv_path):
        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                rows.append({k.strip(): v.strip() if isinstance(v, str) else v for k, v in row.items()})

    if not rows:
        print("No OAPT data found")
        return {}

    # Build per-IATA report lists
    reports = {}
    concerns_set = set()
    types_set = set()

    for row in rows:
        raw_city = (row.get('IncidentCity') or '').strip().upper()
        # Convert ICAO to IATA (e.g., CYWG → YWG, KIAH → IAH)
        iata = ''
        if raw_city and raw_city != 'ENRTE':
            if len(raw_city) == 4 and raw_city[0] in 'CDEFGHKLMNPRSTVW':
                iata = raw_city[1:]
            elif len(raw_city) == 3:
                iata = raw_city
        if not iata:
            raw_dep = (row.get('DepartureCity') or '').strip().upper()
            if len(raw_dep) == 4 and raw_dep[0] in 'CDEFGHKLMNPRSTVW':
                iata = raw_dep[1:]
            elif len(raw_dep) == 3:
                iata = raw_dep
        if not iata:
            raw_arr = (row.get('DestinationCity') or '').strip().upper()
            if len(raw_arr) == 4 and raw_arr[0] in 'CDEFGHKLMNPRSTVW':
                iata = raw_arr[1:]
            elif len(raw_arr) == 3:
                iata = raw_arr
        if not iata:
            continue

        report = {
            'r': str(row.get('ReportNo', '')),
            'c': str(row.get('ImportDate', '')),
            'i': str(row.get('ImportedBy', '')),
            'o': str(row.get('OccNo', '')),
            'd': str(row.get('ReportDesc', '') or row.get('OccDesc', ''))[:800],
            'od': str(row.get('OccDesc', '') or '')[:1200],
            'n': [],
            't': str(row.get('OAPT_Concerns', '') or row.get('OccType', '')),
            'l': str(row.get('DateClosed', '')),
            'b': str(row.get('ClosedBy', '')),
        }

        # Parse concerns (pipe-separated)
        concerns_raw = str(row.get('OAPT_Concerns', ''))
        if concerns_raw and concerns_raw != 'None':
            parts = [c.strip() for c in concerns_raw.split('|') if c.strip()]
            if parts:
                report['n'] = parts
                for c in parts:
                    concerns_set.add(c)

        if report['t']:
            types_set.add(report['t'])

        if iata not in reports:
            reports[iata] = []
        reports[iata].append(report)

    # Write oapt_reports.js
    output = 'var OAPT_REPORTS = ' + json.dumps({
        'reports': reports,
        'types': sorted(types_set),
        'concerns': sorted(concerns_set),
    }, ensure_ascii=False) + ';\n'

    with open(os.path.join(ROOT, 'data', 'oapt_reports.js'), 'w', encoding='utf-8') as f:
        f.write(output)

    print(f"OAPT: {len(reports)} stations, {sum(len(v) for v in reports.values())} reports")
    return {'reports': reports, 'types': sorted(types_set), 'concerns': sorted(concerns_set)}


def process_crs(oapt_data=None):
    """Process CRS data and merge with OAPT descriptions."""
    filepath = find_data_file('Group Safety CRS Report')
    if not filepath:
        print("CRS file not found")
        return

    print(f"CRS source file: {filepath}")

    rows = read_data_rows(filepath)
    if not rows:
        print("CRS file empty")
        return

    # Count FDWTR before dedup
    fdwtr_before = sum(1 for r in rows if (r.get('OccType') or '').strip() == 'FDWTR')
    print(f"CRS rows: {len(rows)}, FDWTR before dedup: {fdwtr_before}")

    # Normalize headers to standard names
    CRS_HEADER_MAP = {
        'OccID': 'OccNo', 'OccurredOn': 'Occ Date', 'Flight_Number': 'Flight No',
        'Tail': 'Tail No', 'SeatConfig': 'Seat Config', 'Aircraft_Type': 'AC Model',
        'Incident_City': 'Incident City', 'Dep_City': 'Dep City', 'Arr_City': 'Arr City',
        'DescriptorSet': 'Descriptor Set', 'Level_1': 'Level 1', 'Level_2': 'Level 2',
        'Level_3': 'Level 3', 'Risk_Level': 'Risk Level', 'Cargo_AWB': 'Cargo AWB',
        'Special_Handling_Code__SHC_': 'Special Handling Code',
        'HFACS_Lvl_1': 'HFACS Lvl 1', 'HFACS_Lvl_2': 'HFACS Lvl 2',
        'HFACS_Lvl_3': 'HFACS Lvl 3', 'Phase_of_Operation': 'Phase of Operation',
        'Type_of_Damage': 'Type of Damage',
    }
    rows = [{CRS_HEADER_MAP.get(k, k): v for k, v in r.items()} for r in rows]

    # Build OAPT OccNo -> ReportDesc lookup (also match by base number ignoring year suffix)
    oapt_desc = {}
    oapt_reports_lookup = {}
    oapt_by_base = {}  # base_occ -> list of OAPT report entries
    if oapt_data and 'reports' in oapt_data:
        for iata, reports in oapt_data['reports'].items():
            for r in reports:
                occ_no = r.get('o', '')
                if occ_no:
                    if occ_no not in oapt_desc and r.get('d'):
                        oapt_desc[occ_no] = r['d']
                    if occ_no not in oapt_desc and r.get('od'):
                        oapt_desc[occ_no] = r['od']
                    if occ_no not in oapt_reports_lookup:
                        oapt_reports_lookup[occ_no] = r
                    base = re.sub(r'-\d+$', '', occ_no)
                    if base not in oapt_by_base:
                        oapt_by_base[base] = []
                    oapt_by_base[base].append(r)

    records = []

    def g(row, field):
        v = row.get(field, '')
        if v is None:
            return ''
        return str(v).strip() if not isinstance(v, (int, float)) else str(v).strip()

    # Group raw rows by (occ_no, occ_type). One occurrence can have multiple
    # descriptor-set rows (DescriptorSet/Level_1/Level_2/...), so we keep every
    # distinct descriptor row and only merge the OAPT description once per group.
    groups = {}
    for row in rows:
        occ_no = g(row, 'OccNo')
        if not occ_no:
            continue
        occ_type = g(row, 'OccType')
        if not occ_type or not re.match(r'^[A-Z][A-Z0-9\-]+$', occ_type):
            continue
        groups.setdefault((occ_no, occ_type), []).append(row)

    for (occ_no, occ_type), group in groups.items():
        row0 = group[0]

        city = g(row0, 'Incident City').upper()
        dep = g(row0, 'Dep City').upper()
        arr = g(row0, 'Arr City').upper()

        # Get description from CRS or OAPT (use OccDesc if ReportDesc is too short)
        crs_desc = g(row0, 'Description')
        oapt_desc_val = oapt_desc.get(occ_no, '')
        # Also grab OccDesc for search
        oapt_occ_desc = ''
        if occ_no in oapt_reports_lookup:
            oapt_occ_desc = str(oapt_reports_lookup[occ_no].get('od', '') or '')
        base = re.sub(r'-\d+$', '', occ_no)
        if not oapt_occ_desc and base in oapt_by_base:
            for r in oapt_by_base[base]:
                od = str(r.get('od', '') or '')
                if len(od) > len(oapt_occ_desc):
                    oapt_occ_desc = od
        rd = oapt_occ_desc if oapt_occ_desc and len(oapt_occ_desc) >= len(oapt_desc_val) else (oapt_desc_val if oapt_desc_val else crs_desc)
        # Build full searchable text — prioritize OccDesc (richest for search)
        search_text = ' '.join(filter(None, [oapt_occ_desc, oapt_desc_val, crs_desc]))[:1000]

        # Get concerns and report count from OAPT (match by base OccNo)
        concerns = []
        base = re.sub(r'-\d+$', '', occ_no)
        report_count = 0
        if base in oapt_by_base:
            report_count = len(oapt_by_base[base])
            for r in oapt_by_base[base]:
                for c in r.get('n', []):
                    if c not in concerns:
                        concerns.append(c)
        elif occ_no in oapt_reports_lookup:
            report_count = 1
            concerns = oapt_reports_lookup[occ_no].get('n', [])

        # Get region from OAPT or derive from city
        region = ''
        if occ_no in oapt_reports_lookup:
            # We'll assign region based on IATA→station mapping later
            pass

        # Determine region from city code (ICAO prefix)
        region = get_region_from_icao(city) or get_region_from_icao(dep) or get_region_from_icao(arr) or 'Western Canada & Mexico'

        # Collect every distinct descriptor-set row for this occurrence
        ds = []
        seen_ds = set()
        for row in group:
            dd = g(row, 'Descriptor Set')
            if not dd:
                continue
            entry = {
                'd': dd,
                'l1': g(row, 'Level 1'),
                'l2': g(row, 'Level 2'),
                'l3': g(row, 'Level 3'),
                'rl': g(row, 'Risk Level'),
                'h1': g(row, 'HFACS Lvl 1'),
                'h2': g(row, 'HFACS Lvl 2'),
                'h3': g(row, 'HFACS Lvl 3'),
            }
            ds_key = (entry['d'], entry['l1'], entry['l2'], entry['rl'], entry['h1'])
            if ds_key in seen_ds:
                continue
            seen_ds.add(ds_key)
            ds.append(entry)

        record = {
            'o': occ_no,
            't': occ_type,
            'd': ds[0]['d'] if ds else '',
            'ds': [{k: v for k, v in e.items() if v} for e in ds],
            'c': city,
            'r': region,
            'n': concerns,
            'rc': report_count,
            'f': g(row0, 'Flight No'),
            'dt': str(g(row0, 'Occ Date')),
            'dep': dep,
            'arr': arr,
            'rd': rd[:1000] if rd else '',
            'st': search_text,
            'al': g(row0, 'Airline'),
            'ac': g(row0, 'AC Model'),
        }
        records.append(record)

    fdwtr_after = sum(1 for r in records if r['t'] == 'FDWTR')
    print(f"CRS records: {len(records)}, FDWTR: {fdwtr_after}")

    # Build IATA→region mapping from OAPT data using ICAO→IATA conversion
    iata_regions = {}
    if oapt_data and 'reports' in oapt_data:
        for iata in oapt_data['reports']:
            # Map IATA back to ICAO to use REGION_ICAO
            icao_candidates = ['C' + iata, 'K' + iata]
            region = None
            for icao in icao_candidates:
                region = get_region_from_icao(icao)
                if region:
                    break
            iata_regions[iata] = region or 'US & Caribbean'

    # Refine region assignment using OAPT lookup (try base match too)
    # Only override if the ICAO-based region isn't already a good match
    icao_regions = set(REGION_ICAO.keys())

    # Precompute base→occ mapping to avoid O(n*m) scan
    base_to_oapt_occ = {}
    for oapt_occ in oapt_reports_lookup:
        base_key = re.sub(r'-\d+$', '', oapt_occ)
        if base_key not in base_to_oapt_occ:
            base_to_oapt_occ[base_key] = oapt_occ

    # Precompute occ→iata mapping from OAPT data
    occ_to_iata = {}
    if oapt_data and 'reports' in oapt_data:
        for iata, rpt_list in oapt_data['reports'].items():
            for rpt in rpt_list:
                occ_val = rpt.get('o', '')
                if occ_val and occ_val not in occ_to_iata:
                    occ_to_iata[occ_val] = iata

    for rec in records:
        matched_occ = rec['o']
        if matched_occ not in oapt_reports_lookup:
            base = re.sub(r'-\d+$', '', matched_occ)
            matched_occ = base_to_oapt_occ.get(base, matched_occ)
        if matched_occ in oapt_reports_lookup:
            iata = occ_to_iata.get(matched_occ)
            if iata:
                new_region = iata_regions.get(iata)
                # Only override if current region is the fallback
                if rec['r'] == 'Western Canada & Mexico' or rec['r'] not in icao_regions:
                    rec['r'] = new_region or rec['r']

    output = 'var CRS_MERGED_REPORTS = ' + json.dumps(records, ensure_ascii=False) + ';\n'
    with open(os.path.join(ROOT, 'data', 'crs_merged_reports.js'), 'w', encoding='utf-8') as f:
        f.write(output)

    size_kb = os.path.getsize(os.path.join(ROOT, 'data', 'crs_merged_reports.js')) / 1024
    print(f"CRS merged: {len(records)} records, {size_kb:.0f} KB")


def process_risk_profile():
    """Process Airport Safety Risk Profile into station data."""
    filepath = find_data_file('Airport Safety Risk Profile')
    if not filepath:
        print("Risk profile file not found")
        return

    rows = read_data_rows(filepath)
    if not rows:
        print("Risk profile file empty")
        return

    # Group by station + assessment type
    station_data = {}
    for row in rows:
        def g(field):
            v = row.get(field, '')
            if v is None:
                return ''
            return str(v).strip() if not isinstance(v, (int, float)) else v

        station = g('Station')
        if not station or len(str(station)) != 3:
            continue
        station = str(station).upper()

        assessment = g('Assessment Type')
        total = g('Total Risk')

        if station not in station_data:
            station_data[station] = {}

        station_data[station]['asrm'] = g('ASRM')
        station_data[station]['region'] = g('Region') or ''

        # Extract hazards
        hazards = []
        for i in range(1, 6):
            h = g(f'Hazard {i}')
            if h:
                hazards.append(str(h))

        if 'Part C' in str(assessment):
            scores = {}
            part_c_axes = [
                '2.0 Environment & Operating Context',
                '3.1 Standards, Procedures & Training (WestJet)',
                '3.2 Operational Support & Setup (WestJet)',
                '3.3 Safety Management, Coordination & Oversight (WestJet)',
                '4.1 Standards, Procedures & Training (Service Provider)',
                '4.2 Safety Management, Assurance & Risk Monitoring (Service Provider)',
                '4.3 Roles, Accountability & Coordination (Service Provider)',
                '4.4 Staffing, Equipment & Capacity (Service Provider)',
                '5.1 Supervisory Presence & Oversight',
                '5.2 Planning & Priority Management',
                '5.3 Managing Known Issues',
                '6.1 Communication & Team Coordination',
                '6.2 Workload & Time Pressure',
                '7.2 Task Performance',
            ]
            for axis in part_c_axes:
                v = g(axis)
                if v:
                    scores[axis] = str(v)
            station_data[station]['partC'] = {
                'total': total if isinstance(total, (int, float)) else 0,
                'axes': scores,
                'hazards': hazards,
            }

        elif 'Part A' in str(assessment):
            scores = {}
            part_a_axes = [
                '2.1 Airport Authority Safety Management System (SMS) & Safety\nGovernance',
                '2.2.1 Environment', '2.2.2 Tenant Management', '2.2.3 Emergency Response',
                '2.3 Terminal Construction & Temporary Conditions', '3.0 Groundside & Airport Access',
                '4.1 Check-in Layout, Capacity & Passenger Flow', '4.2 Baggage Induction at Check-in',
                '5.1 Baggage System — Sortation & Screening Infrastructure',
                '5.2 Bag Room Operations Environment', '5.3 Arrivals, Transfer & Baggage Reclaim',
                '6.1 Boarding Lounge & Gate Area', '6.2 Gate & Boarding Bridge Operations',
                '6.3 Passenger Transport / Remote Stand Operations',
                '7.1 Stand Layout, Markings & Clearances', '7.2 Parking Guidance — Arrival & Departure',
                '7.3 Stand Services & GSE Staging', '7.4 Vehicle Service Roads & Apron Traffic',
                '8.1 Potable Water Servicing Infrastructure', '8.2 Fueling Infrastructure',
                '8.3 Lavatory & Waste Servicing Infrastructure',
                '8.4 De-icing Infrastructure & Operation Setup',
            ]
            for axis in part_a_axes:
                v = g(axis)
                if v:
                    scores[axis] = str(v)
            station_data[station]['partA'] = {
                'total': total if isinstance(total, (int, float)) else 0,
                'axes': scores,
                'hazards': hazards,
            }

        elif 'Part B' in str(assessment):
            scores = {}
            part_b_axes = [
                '2.1 Safety Governance', '2.2 Quality Assurance', '2.3 Safety Risk Management',
                '3.0 Organizational Strucutre & Staffing', '4.1 Service Provider Training',
                '4.2 Training Differenes & Awareness', '5.1 Procedural Framework & Alignment',
                '5.2 Potable Water Servicing', '5.3  Load Control & Loading',
                '5.4 Emergency Response Procedures', '6.1 GSE Program', '6.2 GSE Maintenance',
            ]
            for axis in part_b_axes:
                v = g(axis)
                if v:
                    scores[axis] = str(v)
            station_data[station]['partB'] = {
                'total': total if isinstance(total, (int, float)) else 0,
                'axes': scores,
                'hazards': hazards,
            }

    # Write as JSON for the app to consume
    output_path = os.path.join(ROOT, 'data', 'risk_profile.json')
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(station_data, f, ensure_ascii=False, indent=2)

    print(f"Risk profile: {len(station_data)} stations")


def process_flight_counts():
    """Process flight count data into flight_counts.js."""
    from collections import defaultdict

    filepath = find_data_file('data (11)')
    if not filepath:
        print("Flight counts file not found")
        return

    rows = read_data_rows(filepath)
    if not rows:
        print("Flight counts file empty")
        return

    station_totals = defaultdict(int)
    station_dates = defaultdict(set)
    all_rows_for_daily = []
    for row in rows:
        station = (row.get('Departure Station Code') or '').strip().upper()
        fc_raw = row.get('Flight Counts')
        date = row.get('Scheduled Departure Date (Local)', '')
        if not station or not fc_raw:
            continue
        try:
            fc = int(float(str(fc_raw)))
        except (ValueError, TypeError):
            continue
        station_totals[station] += fc
        if date:
            dt_str = str(date)[:10]
            station_dates[station].add(dt_str)
            all_rows_for_daily.append((station, dt_str, fc))

    flight_data = {}
    for station in station_totals:
        dates = sorted(station_dates.get(station, set()))
        # Build per-date flight counts
        daily = defaultdict(int)
        for st, dt, fc in all_rows_for_daily:
            if st == station:
                daily[dt] += fc
        flight_data[station] = {
            'total': station_totals[station],
            'dates': len(dates),
            'from': dates[0] if dates else '',
            'to': dates[-1] if dates else '',
            'daily': dict(sorted(daily.items())),
        }

    output = 'var FLIGHT_COUNTS = ' + json.dumps(flight_data, ensure_ascii=False) + ';\n'
    with open(os.path.join(ROOT, 'data', 'flight_counts.js'), 'w', encoding='utf-8') as f:
        f.write(output)

    total_flights = sum(v['total'] for v in flight_data.values())
    print(f"Flight counts: {len(flight_data)} stations, {total_flights:,} total flights")


if __name__ == '__main__':
    print("=== Processing data files ===")
    oapt_data = process_oapt()
    process_crs(oapt_data)
    process_risk_profile()
    process_flight_counts()
    print("=== Done ===")
