#!/usr/bin/env python3
"""Station Risk – development server with upload & auto‑processing pipeline.

Usage:
    python3 server.py          # start on port 9000
    python3 server.py --run    # run pipeline immediately and exit
    python3 server.py --watch  # watch for file changes and auto‑process
"""
import json, os, re, subprocess, sys, http.server, cgi, shutil, glob, time

ROOT = os.path.dirname(os.path.abspath(__file__))
PORT = 9000

# ── file type → base filename mapping ─────────────────────────────────────────
FILE_TYPE_MAP = {
    'profile': 'Airport Safety Risk Profile',
    'crs':      'Group Safety CRS Report',
    'oapt':     'OAPT Occurrence Report',
    'flight':   'data (11)',
}

# ── text label → numeric score (mirrors EXCEL_LABEL_MAP in app.js) ────────────
LABEL_SCORE = {'low': 1, 'medium': 2, 'high': 3, 'very high': 4, 'n/a': None}

# ── axis name → ID maps (mirrors EXCEL_AXIS_MAP_{A,B,C} in app.js) ───────────
def _an(name):
    return name.strip().lower().replace('\n', '').replace(' ', '')

EXCEL_AXIS_MAP_A = {_an(k): v for k, v in {
    '2.1 Airport Authority Safety Management System (SMS) & Safety\nGovernance': 'aa-sms-governance',
    '2.2.1 Environment': 'environment',
    '2.2.2 Tenant Management': 'tenant-mgmt',
    '2.2.3 Emergency Response': 'emergency-response',
    '2.3 Terminal Construction & Temporary Conditions': 'terminal-construction',
    '3.0 Groundside & Airport Access': 'groundside-access',
    '4.1 Check-in Layout, Capacity & Passenger Flow': 'checkin-layout',
    '4.2 Baggage Induction at Check-in': 'baggage-induction',
    '5.1 Baggage System — Sortation & Screening Infrastructure': 'baggage-sortation',
    '5.2 Bag Room Operations Environment': 'bag-room',
    '5.3 Arrivals, Transfer & Baggage Reclaim': 'arrivals-reclaim',
    '6.1 Boarding Lounge & Gate Area': 'boarding-lounge',
    '6.2 Gate & Boarding Bridge Operations': 'gate-operations',
    '6.3 Passenger Transport / Remote Stand Operations': 'remote-stand',
    '7.1 Stand Layout, Markings & Clearances': 'stand-layout',
    '7.2 Parking Guidance — Arrival & Departure': 'parking-guidance',
    '7.3 Stand Services & GSE Staging': 'gse-staging',
    '7.4 Vehicle Service Roads & Apron Traffic': 'vehicle-service-roads',
    '8.1 Potable Water Servicing Infrastructure': 'potable-water',
    '8.2 Fueling Infrastructure': 'fueling',
    '8.3 Lavatory & Waste Servicing Infrastructure': 'lavatory-waste',
    '8.4 De-icing Infrastructure & Operation Setup': 'deicing',
}.items()}

EXCEL_AXIS_MAP_B = {_an(k): v for k, v in {
    '2.1 Safety Governance': 'sp-sms-governance',
    '3.0 Organizational Strucutre & Staffing': 'staffing-structure',
    '3.0 Organizational Structure & Staffing': 'staffing-structure',
    '4.1 Service Provider Training': 'training-competency',
    '5.1 Procedural Framework & Alignment': 'procedures-alignment',
    '5.2 Potable Water Servicing': 'potable-water-sp',
    '5.3  Load Control & Loading': 'load-control',
    '5.3 Load Control & Loading': 'load-control',
    '5.4 Emergency Response Procedures': 'emergency-response-sp',
    '6.1 GSE Program': 'gse-management',
}.items()}

EXCEL_AXIS_MAP_C = {_an(k): v for k, v in {
    '2.0 Environment & Operating Context': 'environment-context',
    '3.1 Standards, Procedures & Training (WestJet)': 'wj-standards',
    '3.2 Operational Support & Setup (WestJet)': 'wj-operational-support',
    '3.3 Safety Management, Coordination & Oversight (WestJet)': 'wj-safety-mgmt',
    '4.1 Standards, Procedures & Training (Service Provider)': 'sp-standards',
    '4.2 Safety Management, Assurance & Risk Monitoring (Service Provider)': 'sp-safety-assurance',
    '4.3 Roles, Accountability & Coordination (Service Provider)': 'sp-roles-coordination',
    '4.4 Staffing, Equipment & Capacity (Service Provider)': 'sp-staffing-equipment',
    '5.1 Supervisory Presence & Oversight': 'supervisory-presence',
    '5.2 Planning & Priority Management': 'planning-priority',
    '5.3 Managing Known Issues': 'managing-issues',
    '6.1 Communication & Team Coordination': 'communication-coordination',
    '6.2 Workload & Time Pressure': 'workload-pressure',
    '7.2 Task Performance': 'task-performance',
}.items()}

# ── direct risk profile processing (avoids process_data.py timeout on big files) ─

PART_C_AXES_NAMES = [
    '2.0 Environment & Operating Context',
    '3.1 Standards, Procedures & Training (WestJet)',
    '3.2 Operational Support & Setup (WestJet)',
    '3.3 Safety Management, Coordination & Oversight (WestJet)',
    '4.1 Standards, Procedures & Training (Service Provider)',
    '4.2 Safety Management, Assurance & Risk Monitoring (Service Provider)',
    '4.3 Roles, Accountability & Coordination (Service Provider)',
    '4.4 Staffing, Equipment & Capacity (Service Provider)',
    '5.1 Supervisory Presence & Oversight',
    '5.2 Planning & Priority Management',
    '5.3 Managing Known Issues',
    '6.1 Communication & Team Coordination',
    '6.2 Workload & Time Pressure',
    '7.2 Task Performance',
]
PART_A_AXES_NAMES = [
    '2.1 Airport Authority Safety Management System (SMS) & Safety\nGovernance',
    '2.2.1 Environment', '2.2.2 Tenant Management', '2.2.3 Emergency Response',
    '2.3 Terminal Construction & Temporary Conditions', '3.0 Groundside & Airport Access',
    '4.1 Check-in Layout, Capacity & Passenger Flow', '4.2 Baggage Induction at Check-in',
    '5.1 Baggage System — Sortation & Screening Infrastructure',
    '5.2 Bag Room Operations Environment', '5.3 Arrivals, Transfer & Baggage Reclaim',
    '6.1 Boarding Lounge & Gate Area', '6.2 Gate & Boarding Bridge Operations',
    '6.3 Passenger Transport / Remote Stand Operations',
    '7.1 Stand Layout, Markings & Clearances', '7.2 Parking Guidance — Arrival & Departure',
    '7.3 Stand Services & GSE Staging', '7.4 Vehicle Service Roads & Apron Traffic',
    '8.1 Potable Water Servicing Infrastructure', '8.2 Fueling Infrastructure',
    '8.3 Lavatory & Waste Servicing Infrastructure', '8.4 De-icing Infrastructure & Operation Setup',
]
PART_B_AXES_NAMES = [
    '2.1 Safety Governance', '2.2 Quality Assurance', '2.3 Safety Risk Management',
    '3.0 Organizational Strucutre & Staffing', '4.1 Service Provider Training',
    '4.2 Training Differenes & Awareness', '5.1 Procedural Framework & Alignment',
    '5.2 Potable Water Servicing', '5.3  Load Control & Loading',
    '5.4 Emergency Response Procedures', '6.1 GSE Program', '6.2 GSE Maintenance',
]


def process_risk_profile_direct():
    """Read Airport Safety Risk Profile.xlsx directly and write risk_profile.json."""
    filepath = find_data_file('Airport Safety Risk Profile')
    if not filepath:
        print("Risk profile file not found, skipping")
        return

    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header = [str(h).strip() if h else '' for h in all_rows[0]]
    rows = []
    for r in all_rows[1:]:
        row = {}
        for i, h in enumerate(header):
            if i < len(r) and r[i] is not None:
                val = r[i]
                if isinstance(val, str):
                    val = val.strip()
                elif hasattr(val, 'strftime'):
                    val = val.strftime('%Y-%m-%d')
                else:
                    val = str(val).strip() if val else ''
                row[h] = val
        if any(row.values()):
            rows.append(row)

    station_data = {}
    for row in rows:
        station = str(row.get('Station', '')).strip().upper()
        if len(station) != 3:
            continue
        assessment = str(row.get('Assessment Type', ''))
        total = row.get('Total Risk', 0)

        if station not in station_data:
            station_data[station] = {'asrm': row.get('ASRM', ''), 'region': row.get('Region', '') or ''}

        hazards = []
        for i in range(1, 6):
            h = row.get(f'Hazard {i}')
            if h:
                hazards.append(str(h))

        if 'Part C' in assessment:
            scores = {a: str(row.get(a, '')) for a in PART_C_AXES_NAMES if row.get(a)}
            station_data[station]['partC'] = {
                'total': total if isinstance(total, (int, float)) else 0,
                'axes': scores, 'hazards': hazards,
            }
        elif 'Part A' in assessment:
            scores = {a: str(row.get(a, '')) for a in PART_A_AXES_NAMES if row.get(a)}
            station_data[station]['partA'] = {
                'total': total if isinstance(total, (int, float)) else 0,
                'axes': scores, 'hazards': hazards,
            }
        elif 'Part B' in assessment:
            scores = {a: str(row.get(a, '')) for a in PART_B_AXES_NAMES if row.get(a)}
            station_data[station]['partB'] = {
                'total': total if isinstance(total, (int, float)) else 0,
                'axes': scores, 'hazards': hazards,
            }

    json_path = os.path.join(ROOT, 'data', 'risk_profile.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(station_data, f, ensure_ascii=False, indent=2)
    print(f"Risk profile: {len(station_data)} stations")


# ── helpers ───────────────────────────────────────────────────────────────────

def find_data_file(basename):
    for ext in ['.xlsx', '.xls', '.csv']:
        p = os.path.join(ROOT, basename + ext)
        if os.path.exists(p):
            return p
    return None


def excel_label_to_score(lbl):
    if not lbl:
        return None
    key = str(lbl).strip().lower()
    return LABEL_SCORE.get(key, None)


def map_risk_axes(axes_obj, axis_map):
    result = {}
    for name, lbl in axes_obj.items():
        key = _an(name)
        aid = axis_map.get(key)
        if aid:
            result[aid] = excel_label_to_score(lbl)
    return result


# ── risk_profile.json → risk_profile.js ───────────────────────────────────────

def convert_risk_profile_json_to_js():
    json_path = os.path.join(ROOT, 'data', 'risk_profile.json')
    js_path = os.path.join(ROOT, 'data', 'risk_profile.js')
    if not os.path.exists(json_path):
        return False
    with open(json_path, encoding='utf-8') as f:
        data = json.load(f)
    with open(js_path, 'w', encoding='utf-8') as f:
        f.write('var RISK_PROFILE = ' + json.dumps(data, ensure_ascii=False, indent=2) + ';\n')
    size = os.path.getsize(js_path) / 1024
    print(f"risk_profile.js: {len(data)} stations, {size:.0f} KB")
    return True


# ── risk_profile.json → seed_data.js merge ────────────────────────────────────

PART_AXES = {
    'partA': ['aa-sms-governance', 'environment', 'tenant-mgmt', 'emergency-response',
              'terminal-construction', 'groundside-access', 'checkin-layout',
              'baggage-induction', 'baggage-sortation', 'bag-room', 'arrivals-reclaim',
              'boarding-lounge', 'gate-operations', 'remote-stand', 'stand-layout',
              'parking-guidance', 'gse-staging', 'vehicle-service-roads', 'potable-water',
              'fueling', 'lavatory-waste', 'deicing'],
    'partB': ['sp-sms-governance', 'staffing-structure', 'training-competency',
              'procedures-alignment', 'potable-water-sp', 'load-control',
              'emergency-response-sp', 'gse-management'],
    'partC': ['environment-context', 'wj-standards', 'wj-operational-support',
              'wj-safety-mgmt', 'sp-standards', 'sp-safety-assurance', 'sp-roles-coordination',
              'sp-staffing-equipment', 'supervisory-presence', 'planning-priority',
              'managing-issues', 'communication-coordination', 'workload-pressure',
              'task-performance'],
}


def parse_seed_data_js():
    path = os.path.join(ROOT, 'data', 'seed_data.js')
    if not os.path.exists(path):
        return None, {}, ''
    with open(path, encoding='utf-8') as f:
        content = f.read()
    version_match = re.search(r'var\s+SEED_VERSION\s*=\s*"([^"]*)"', content)
    version = version_match.group(1) if version_match else ''
    data_match = re.search(r'var\s+SEED_DATA\s*=\s*(\{.*?\});\s*$', content, re.DOTALL)
    if not data_match:
        return version, {}, content
    try:
        seed = json.loads(data_match.group(1))
    except json.JSONDecodeError:
        return version, {}, content
    return version, seed, content


def write_seed_data_js(version, seed):
    # Bump version
    today = time.strftime('%Y-%m-%d')
    v_match = re.match(r'(\d{4}-\d{2}-\d{2})-v(\d+)', version)
    if v_match and v_match.group(1) == today:
        next_v = int(v_match.group(2)) + 1
    else:
        next_v = 1
    new_version = f"{today}-v{next_v}"
    output = f'var SEED_VERSION = "{new_version}";\n'
    output += 'var SEED_DATA = ' + json.dumps(seed, ensure_ascii=False, separators=(',', ':')) + ';\n'
    path = os.path.join(ROOT, 'data', 'seed_data.js')
    with open(path, 'w', encoding='utf-8') as f:
        f.write(output)
    size = os.path.getsize(path) / 1024
    print(f"seed_data.js: {len(seed)} stations v{new_version}, {size:.0f} KB")


def ensure_part_b_target(s):
    pb = s.get('partB')
    if isinstance(pb, list):
        if not pb:
            pb.append({'serviceProvider': '', 'function': '', 'status': 'not-started', 'date': '', 'scores': {}, 'notes': {}})
        return pb[0]
    elif isinstance(pb, dict):
        return pb
    else:
        s['partB'] = {'status': 'not-started', 'date': '', 'scores': {}, 'notes': {}}
        return s['partB']


def merge_risk_into_seed():
    json_path = os.path.join(ROOT, 'data', 'risk_profile.json')
    if not os.path.exists(json_path):
        print("risk_profile.json not found, skipping seed merge")
        return False

    with open(json_path, encoding='utf-8') as f:
        risk = json.load(f)

    version, seed, _ = parse_seed_data_js()
    if not seed:
        print("No existing seed_data.js found, creating new one")
        seed = {}
        version = ''

    changed = 0
    for iata, entry in risk.items():
        # Only process stations that have actual Excel data
        excel_has_data = any(
            ep and ep.get('axes') and len(ep['axes']) > 0
            for pk in ('partA', 'partB', 'partC')
            if (ep := entry.get(pk))
        )
        if not excel_has_data:
            continue

        if iata not in seed:
            s = {
                'iataCode': iata, 'name': '', 'airportName': '', 'location': '',
                'region': entry.get('region', ''),
                'partA': {'status': 'not-started', 'date': '', 'scores': {}, 'notes': {}},
                'partB': {'status': 'not-started', 'date': '', 'scores': {}, 'notes': {}},
                'partC': {'status': 'not-started', 'date': '', 'scores': {}, 'notes': {}},
                'history': [],
                'operationalData': {'flightNumbers': '', 'exposure': '', 'qci': '',
                                    'auditFindings': '', 'incidentTrends': []},
            }
            seed[iata] = s
        else:
            s = seed[iata]

        s['region'] = entry.get('region', s.get('region', ''))

        for part_key, axis_map, axes_list in [
            ('partA', EXCEL_AXIS_MAP_A, PART_AXES['partA']),
            ('partB', EXCEL_AXIS_MAP_B, PART_AXES['partB']),
            ('partC', EXCEL_AXIS_MAP_C, PART_AXES['partC']),
        ]:
            excel_part = entry.get(part_key)
            if not excel_part or not excel_part.get('axes'):
                continue

            mapped = map_risk_axes(excel_part['axes'], axis_map)
            if not mapped:
                continue

            if part_key == 'partB':
                target = ensure_part_b_target(s)
            else:
                target = s[part_key]

            if 'scores' not in target:
                target['scores'] = {}
            for aid in axes_list:
                if aid in mapped:
                    old_val = target['scores'].get(aid)
                    new_val = mapped[aid]
                    if old_val != new_val:
                        changed += 1
                    target['scores'][aid] = new_val

            if any(v is not None for v in mapped.values()):
                if target.get('status') != 'complete':
                    changed += 1
                target['status'] = 'complete'

    write_seed_data_js(version, seed)
    print(f"Merged risk data: {changed} score changes across {len(risk)} stations")
    return True


# ── pipeline ──────────────────────────────────────────────────────────────────

def run_pipeline(saved_files=None):
    print("=== Processing data files ===")
    output = ""

    # 1. Run our own risk profile processing (fast, targeted)
    process_risk_profile_direct()

    # 2. Convert risk_profile.json → risk_profile.js
    convert_risk_profile_json_to_js()

    # 3. Merge risk profile into seed_data.js
    merge_risk_into_seed()

    # 4. Try process_data.py for OAPT/CRS/flight (with timeout)
    proc_path = os.path.join(ROOT, 'process_data.py')
    if os.path.exists(proc_path):
        try:
            result = subprocess.run(
                [sys.executable, proc_path],
                capture_output=True, text=True, cwd=ROOT,
                timeout=120
            )
            extra = result.stdout + '\n' + result.stderr
            output += extra
            print(extra)
            if result.returncode != 0:
                print("WARNING: process_data.py had errors (non-fatal)")
        except subprocess.TimeoutExpired:
            print("process_data.py timed out (large files) – risk profile already processed")
        except Exception as e:
            print(f"process_data.py error: {e}")
    else:
        print("process_data.py not found, skipping")

    print("=== Done ===")
    return True, output


# ── HTTP server ───────────────────────────────────────────────────────────────

class UploadHandler(http.server.SimpleHTTPRequestHandler):

    def do_GET(self):
        if self.path == '/api/health':
            self.send_json({'status': 'ok'})
            return
        return super().do_GET()

    def do_POST(self):
        if self.path != '/api/upload':
            self.send_error(404)
            return

        content_type = self.headers.get('Content-Type', '')
        if 'multipart/form-data' not in content_type:
            self.send_json({'success': False, 'error': 'Expected multipart/form-data'})
            return

        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={
                'REQUEST_METHOD': 'POST',
                'CONTENT_TYPE': content_type,
            }
        )

        saved_files = []
        for field_name, basename in FILE_TYPE_MAP.items():
            item = form.get(field_name)
            if item and item.file:
                ext = os.path.splitext(item.filename or '')[1] or '.xlsx'
                dest = basename + ext
                dest_path = os.path.join(ROOT, dest)
                with open(dest_path, 'wb') as f:
                    shutil.copyfileobj(item.file, f)
                saved_files.append({'original': item.filename, 'saved': dest})
                print(f"Saved: {item.filename} → {dest}")

        if not saved_files:
            self.send_json({'success': False, 'error': 'No valid files uploaded'})
            return

        success, output = run_pipeline(saved_files)
        self.send_json({
            'success': success,
            'saved': saved_files,
            'output': output,
            'error': None if success else 'Processing failed – see output for details',
        })

    def send_json(self, obj):
        body = json.dumps(obj).encode('utf-8')
        self.send_response(200)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Content-Length', str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format, *args):
        print(f"[{self.log_date_time_string()}] {format % args}")


# ── file watcher ──────────────────────────────────────────────────────────────

def watch_and_process():
    watched = {}
    for basename in FILE_TYPE_MAP.values():
        for ext in ['.xlsx', '.csv']:
            p = os.path.join(ROOT, basename + ext)
            if os.path.exists(p):
                watched[p] = os.path.getmtime(p)

    print(f"Watching {len(watched)} file(s) for changes…")
    try:
        while True:
            time.sleep(2)
            changed = []
            for fp in list(watched):
                if os.path.exists(fp):
                    mtime = os.path.getmtime(fp)
                    if mtime != watched[fp]:
                        watched[fp] = mtime
                        changed.append(os.path.basename(fp))
                else:
                    watched.pop(fp, None)
            # Also check for new files
            for basename in FILE_TYPE_MAP.values():
                for ext in ['.xlsx', '.csv']:
                    p = os.path.join(ROOT, basename + ext)
                    if os.path.exists(p) and p not in watched:
                        watched[p] = os.path.getmtime(p)
                        changed.append(os.path.basename(p))

            if changed:
                print(f"\nDetected changes: {', '.join(changed)}")
                run_pipeline()
    except KeyboardInterrupt:
        print("\nWatcher stopped")


# ── entry point ───────────────────────────────────────────────────────────────

if __name__ == '__main__':
    if '--run' in sys.argv:
        run_pipeline()
    elif '--watch' in sys.argv:
        watch_and_process()
    else:
        os.chdir(ROOT)
        server = http.server.HTTPServer(('0.0.0.0', PORT), UploadHandler)
        print(f"Serving at http://localhost:{PORT}")
        print(f"Upload endpoint: POST http://localhost:{PORT}/api/upload")
        print(f"Open http://localhost:{PORT} in your browser")
        try:
            server.serve_forever()
        except KeyboardInterrupt:
            print("\nServer stopped")
